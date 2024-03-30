import pandas as pd
import openpyxl 

# the output of this file will be an xlsx file in this format:
#     | C1 | C2
#  C1 | 0 | 1
#  C2 | 0 | 0

# input file (example from google forms xlsx export)

xlsx_input_folder = 'input/'
xlsx_output_folder = 'generated_format/'
xlsx_input_file = 'nc.xlsx'
xlsx_input_path = xlsx_input_folder+xlsx_input_file
xlsx_output_path = xlsx_output_folder+xlsx_input_file

# read by default 1st sheet of an excel file
dataframe1 = pd.read_excel(xlsx_input_path)
# Load headers
headers = list(dataframe1.columns.values)
# Exclude headers that are not important
ex = ['Időbélyeg']

for e in ex:
    headers.remove(e)

# build array from header information
l = {};
for header in headers:
    l[header]=dataframe1.groupby([header])

# occurance of question
l2 = {}
# answer rows for question 
l3 = {}

# how many rows exist
rowsize = len(dataframe1.index)
#print(rowsize)

for key, value in l.items():
    # answer array for question
    # [array(['A'], dtype=object), array(['B'], dtype=object)]
    u = list(value[key].unique())

    # count occurance of that said answer for the question
    v = value[key].value_counts()
    

    unikey={}
    
    for uni in u:
        if key in l3:
            l3[str(key)][str(uni[0])] = dataframe1.index[dataframe1[key] == uni[0]].tolist()
        else:
            l3[str(key)] = {str(uni[0]): dataframe1.index[dataframe1[key] == uni[0]].tolist()}

        unikey[str(uni[0])] = int(v[uni])/rowsize

    l2[key] = unikey

wb = openpyxl.Workbook() 
ws = wb.active 
rkerdes = 1
rvalasz = 1

r = rkerdes+rvalasz

def getConnceptLabel(r,label):
    return str(label)+str(r-1)

for i in range(len(dataframe1.index)):
    ws.cell(row = r, column = 1).value = getConnceptLabel(r,"U")
    ws.cell(row = 1, column = r).value = getConnceptLabel(r,"U")
    rkerdes+=1
    r = rkerdes+rvalasz

for row in l2.keys(): 
    ws.cell(row = r, column = 1).value = getConnceptLabel(r,"K")
    ws.cell(row = 1, column = r).value = getConnceptLabel(r,"K")
    
    # relation to it self should be null | önmagába nem esik vissza
    ws.cell(row = r, column = r).value = 0
    rparent = rvalasz+rkerdes
    rkerdes+=1
    r = rkerdes+rvalasz
    
    for value in l2[row]: 
        ws.cell(row = r, column = 1).value = getConnceptLabel(r,"V")
        ws.cell(row = 1, column = r).value = getConnceptLabel(r,"V")
        #print(value)
        for con in l3[row][value]:
            #ws.cell(row = r, column = con+2).value = -1
            ws.cell(row = con+2, column = r).value = l2[row][value]
            
        if r!=rparent:
            ws.cell(row = r, column = rparent).value = 1
            # ws.cell(row =  rparent, column =r).value = 1
        rvalasz+=1
        r = rkerdes+rvalasz
        
wb.save(xlsx_output_path) 
print ("OK")
# We do not know the optimal value of E
# for the reduction we can choose a random value in the range of : ] 0,1 [ 
# or we can do a test clustering run to calculate it.
# TODO
# multithreaded python reduction cluster algorithm

import zlib
import json

matrix_folder = 'zlib_compressed_cluster_build_format/'
matrix_out_folder = 'zlib_compressed_cluster_output/'
matrix_file = 'ajax.php'

# todo change to API url
matrixpath = matrix_folder+matrix_file
with open(matrixpath, 'rb') as f:
    decompressed_data = zlib.decompress(f.read())
y = json.loads(decompressed_data)
initial_conncept = y["init_state"]
n = len(initial_conncept);
W = y["connection_matrix"]

E=0.0015

print("Matrix loaded : "+matrixpath)
print("E value : "+str(E))
print("Matrix size : "+str(n)+"x"+str(n))


def save_json_compressed(data, file_path):
    """
    Save data (converted to JSON) compressed with zlib into a file within a temporary folder.

    Parameters:
    - data: The data to be saved (should be JSON-serializable).
    - filename: The filename to save the compressed JSON data.

    Returns:
    - The path to the saved file.
    """
    try:
        # Convert data to JSON
        json_data = json.dumps(data)

        # Compress the JSON data with zlib
        compressed_data = zlib.compress(json_data.encode())

        # Write the compressed data to a file
        with open(file_path, 'wb') as file:
            file.write(compressed_data)

        return file_path
    except Exception as e:
        print("An error occurred:", e)
        return None


def isNearC(i, j, E):
    sum = 0
    for k in range(n):
        if k != i and k != j:
            #print("Line i",i,"j",j,"k",k," w1",(W[i][k] - W[j][k]),"w2",(W[k][i] - W[k][j]))
            sum += (W[i][k] - W[j][k]) ** 2
            sum += (W[k][i] - W[k][j]) ** 2

    #print(sum / ((n - 2) * 8))
    if (sum / ((n - 2) * 8) < E):
        return True
    else:
        return False


def buildCluster(initial_conncept, E):
    K = {initial_conncept:1}
    currentK = initial_conncept
    for i in range(n):
        if i != initial_conncept:
            member = True
            while member:
                j = currentK
                member = isNearC(j, i, E)
                if member:
                    if i not in K.keys():
                        K[i]=1
                        currentK=i
                    else:
                        member=False
    
    save_json_compressed(K,matrix_out_folder+str(initial_conncept)+".tmp")
        
    return True;

import multiprocessing
from functools import partial
from datetime import datetime

def worker(task_queue, result_queue, task_complete_event):
    try:
        while True:
            task = task_queue.get()
            if task is None:
                print("Worker process terminating")
                break
            # Perform the task
            print("Performing task")
            result = task()
            #result_queue.put(result)
            task_complete_event.set()  # Signal task completion
            print("Performing task-ended-signaled")
    except Exception as e:
        print("Exception occurred:", e)

def task_generator(tasks, task_queue):
    """Generates tasks and puts them into the task queue."""
    for task in tasks:
        task_queue.put(task)

def generate_clusters(i,E):
    """Generate a random number."""
    print("Starting cluster [",str(i),"] Current Time =", datetime.now().strftime("%H:%M:%S"))
    return buildCluster(i,E)

def main():
    # Number of worker processes to create
    num_tasks = n
    num_processes = min(multiprocessing.cpu_count(),num_tasks)

    # Create task queue and result queue
    task_queue = multiprocessing.Queue()
    result_queue = multiprocessing.Queue()

    # Generate tasks
   
    # tasks = [partial( generate_clusters,_,E) for _ in range(num_tasks)]
    tasks = [partial( generate_clusters,_,E) for _ in range(n)]

    # Put tasks into the queue
    task_generator(tasks, task_queue)

    # Event to track task completion
    task_complete_event = multiprocessing.Event()

    # Create worker processes
    processes = []
    
    for _ in range(num_processes):
        p = multiprocessing.Process(target=worker, args=(task_queue, result_queue, task_complete_event))
        p.start()
        processes.append(p)

    # Wait for all tasks to complete
    while True:
        if task_complete_event.wait(timeout=1):  # Check every second
            task_complete_event.clear()
        else:
            # If no tasks completed within timeout, check if all tasks are done
            if task_queue.empty():
                print("task_queue.empty()")
                break

    # Terminate worker processes
    for _ in range(num_processes):
        task_queue.put(None)  # Signal to stop worker

    # Join worker processes
    for p in processes:
        p.join()

    # Collect results
    results = []
    while not result_queue.empty():
        result = result_queue.get()
        results.append(result)

    print("All tasks completed.")
    # Store the JSON data in a file
    with open("res.json", "w") as file:
        json.dump(results, file)
    #print("Results:", results)
    print("Results are written")

if __name__ == "__main__":
    main()

# now we know what are the points that can be put in the same cluster
# this part does calculations for the cluster and generates the modified W with the cluster elements merged 
NEW_W = []
cluster_incoming={}
cluster_incoming_counter={}

cluster_outgoing={}
cluster_outgoing_counter={}

for file in range(n):
    with open(matrix_out_folder+"1.tmp", 'rb') as f:
            data = zlib.decompress(f.read())
            y = json.loads(data)
            clustername = "M"
            # get outgoint connections
            
            for ckey in y:
                #print("get outgoint connections")
                clustername+="_C"+str(int(ckey)+1)

                for i, val in enumerate(W[int(ckey)]):
                    # print("C"+str(int(ckey)+1), y[ckey])
                    if val!=0:
                        # print("C"+str(int(i)+1),val)
                        if int(i) in cluster_outgoing:
                            cluster_outgoing[int(i)]+=val
                            cluster_outgoing_counter[int(i)]+=1
                        else:
                            cluster_outgoing[int(i)]=val
                            cluster_outgoing_counter[int(i)]=1
                    #print("get incoming connections")
                    if W[i][int(ckey)]!=0:
                        #print("C"+str(int(i)+1),val)
                        if int(i) in cluster_incoming:
                            cluster_incoming[int(i)]+=W[i][int(ckey)]
                            cluster_incoming_counter[int(i)]+=1
                        else:
                            cluster_incoming[int(i)]=W[i][int(ckey)]
                            cluster_incoming_counter[int(i)]=1
        
            print("Outgoing connections:")
            for k, v in cluster_outgoing.items():
                print("C"+str(int(k)+1),v/cluster_outgoing_counter[k])
            # get incoming connections
            
            print("Incoming connections:")
            for k, v in cluster_outgoing.items():
                print("C"+str(int(k)+1),v/cluster_outgoing_counter[k])
            
    break
print(clustername)
#print(cluster_incoming)
#print(cluster_incoming_counter)
#print(cluster_outgoing)
#print(cluster_outgoing_counter)